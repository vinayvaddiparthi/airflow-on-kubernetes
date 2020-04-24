import os
import tempfile
from concurrent.futures.thread import ThreadPoolExecutor
from zipfile import BadZipFile

import pandas as pd
from openpyxl import load_workbook
import boto3
from snowflake.sqlalchemy import URL
from sqlalchemy import create_engine

from utils import sf15to18

s3 = boto3.resource("s3")


def _process_excel_file(bucket: str, key: str):
    bucket = s3.Bucket(name=bucket)
    print(f"Processing {bucket}/{key}")

    with tempfile.TemporaryFile() as f:
        bucket.download_fileobj(key, f)

        try:
            workbook = load_workbook(filename=f, data_only=True, read_only=True)
            ws = workbook["Calculation Sheet"]
            calc_sheet = {
                k[0].value: v[0].value for k, v in zip(ws["A1":"A32"], ws["B1":"B32"])
            }
            return calc_sheet
        except BadZipFile:
            pass


if __name__ == "__main__":
    bucket = "tc-workbooks"
    with ThreadPoolExecutor(max_workers=128) as executor:
        futures = [
            executor.submit(_process_excel_file, bucket, obj.key)
            for obj in s3.Bucket(name=bucket).objects.all()
            if (obj.key).lower().endswith(".xlsx") and not "~" in obj.key
        ]

    df = pd.DataFrame([future.result() for future in futures])
    df["Account ID - 18"] = df.apply(lambda row: sf15to18(row["Account ID"]), axis=1)

    engine_ = SnowflakeHook(
        URL(
            account="thinkingcapital.ca-central-1.aws",
            user=os.environ["SNOWFLAKE_USERNAME"],
            password=os.environ["SNOWFLAKE_PASSWORD"],
            database="ANALYTICS_PRODUCTION",
            warehouse="ETL",
            role="SYSADMIN",
        )
    )

    with engine_.begin() as tx:
        accountid = df["Account ID - 18"][0]

        res = tx.execute(
            f"SELECT LS.* "
            f"FROM DBT_SBPORTAL_PII.LOAN_SUMMARY LS "
            f"JOIN DBT_SAS.LOAN_ACCOUNT_SUMMARY LA ON LA.CONTRACT_UUID = LS.CONTRACT_UUID "
            f"WHERE LA.ACCOUNTID='{accountid}' AND LA.CURRENT_CONTRACT=1"
        ).fetchall()
        print(res)
