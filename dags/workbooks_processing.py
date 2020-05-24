import logging
import os
import re
import tempfile
from concurrent.futures.thread import ThreadPoolExecutor
from datetime import timedelta
from pathlib import Path
from typing import Optional, IO

import pandas as pd
import pendulum
from airflow import DAG
from airflow.contrib.hooks.snowflake_hook import SnowflakeHook
from airflow.contrib.kubernetes.secret import Secret
from airflow.contrib.operators.kubernetes_pod_operator import KubernetesPodOperator
from airflow.operators.python_operator import PythonOperator
from openpyxl import load_workbook
import boto3
from slugify import slugify

from utils import sf15to18, random_identifier


def _wrap_sf15to18(id: str) -> Optional[str]:
    try:
        return sf15to18(id)
    except Exception as e:
        logging.error(f"❌ Error converting SFDC ID {id} from 15 to 18 chars: {e}")
        return None


def _process_excel_file(key: str, file: IO[bytes]) -> dict:
    workbook = load_workbook(filename=file, data_only=True, read_only=True)
    ws = workbook["Calculation Sheet"]
    calc_sheet = {
        slugify(str(k[0].value), separator="_"): v[0].value
        for k, v in zip(ws["A1":"A65535"], ws["B1":"B65535"])
        if k and v
    }
    calc_sheet["key"] = key

    return calc_sheet


def _get_and_process_workbook(bucket, obj) -> dict:
    try:
        with tempfile.TemporaryFile() as f:
            bucket.download_fileobj(obj.key, f)
            return _process_excel_file(obj.key, f)
    except Exception as e:
        logging.error(f"❌ Error processing {obj}: {e}")
        return {}


def import_workbooks(
    bucket_name: str,
    snowflake_conn: str,
    destination_schema: str,
    destination_table: str,
    num_threads: int = 4,
):
    # Delete AWS credentials used to upload logs from this context
    try:
        del os.environ["AWS_ACCESS_KEY_ID"]
        del os.environ["AWS_SECRET_ACCESS_KEY"]
    except Exception:  # nosec
        pass

    s3 = boto3.resource("s3")
    bucket = s3.Bucket(name=bucket_name)

    stage_guid = random_identifier()

    with ThreadPoolExecutor(max_workers=num_threads) as executor:
        print(f"⚙️ Processing {bucket}...")
        futures = [
            executor.submit(_get_and_process_workbook, bucket, obj)
            for obj in bucket.objects.all()
            if not re.match("^.*_+[0-9]+.xlsx$", obj.key.lower())
            and "/" not in obj.key
            and not obj.key.lower().startswith("~$")
        ]

        results = [future.result() for future in futures]

    df = pd.DataFrame(results)
    print(f"✔️ Done processing {len(futures)} workbooks")

    df["account_id_18"] = df.apply(
        lambda row: _wrap_sf15to18(row.get("account_id")), axis=1
    )
    print(f"✔️ Done computing 18 characters SFDC object IDs")

    engine_ = SnowflakeHook(snowflake_conn).get_sqlalchemy_engine()
    with engine_.begin() as tx, tempfile.TemporaryDirectory() as path:
        file = Path(path) / random_identifier()
        df.to_json(str(file), orient="records", lines=True)
        print("Dataframe converted to JSON")

        bucket.upload_file(str(file), "_summary.json")
        print(f"📦 {file} uploaded to S3")

        stmts = [
            f"CREATE OR REPLACE TEMPORARY STAGE {destination_schema}.{stage_guid} FILE_FORMAT=(TYPE=JSON)",  # nosec
            f"PUT file://{file} @{destination_schema}.{stage_guid}",  # nosec
            f"CREATE OR REPLACE TRANSIENT TABLE {destination_schema}.{destination_table} AS SELECT $1 AS FIELDS FROM @{destination_schema}.{stage_guid}",  # nosec
        ]

        print("Uploading results to Snowflake ❄️")
        print([tx.execute(stmt).fetchall() for stmt in stmts])


with DAG(
    dag_id="workbooks_processing",
    start_date=pendulum.datetime(
        2020, 5, 3, tzinfo=pendulum.timezone("America/Toronto")
    ),
    schedule_interval="0 6 * * *",
    default_args={"retries": 3, "retry_delay": timedelta(minutes=5)},
) as dag:
    cmdargs = [
        "apt-get update && "
        "apt-get install -y cifs-utils wget && "
        "wget https://dl.min.io/client/mc/release/linux-amd64/mc && "
        "chmod +x mc && mv mc /usr/local/bin && "
        "mkdir ~/.mc &&"
        "cp /secrets/mc-config ~/.mc/config.json && "
        "mkdir /workbooks && "
        'mount -t cifs "//10.10.1.110/Working Files" /workbooks -o credentials=/secrets/auth && '
        "mc mirror --overwrite /workbooks s3/tc-workbooks && "
        "umount /workbooks"
    ]

    grab_workbooks = KubernetesPodOperator(
        task_id="grab_workbooks",
        namespace="airflow",
        image="ubuntu:16.04",
        cmds=["bash", "-cx"],
        arguments=cmdargs,
        secrets=[Secret("volume", "/secrets", "workbooks-secret")],
        name="grab_workbooks",
        is_delete_operator_pod=True,
        privileged=True,
        annotations={
            "iam.amazonaws.com/role": "arn:aws:iam::810110616880:role/"
            "KubernetesAirflowProductionWorkbooksRole"
        },
        executor_config={
            "KubernetesExecutor": {"service_account_name": "airflow-scheduler"}
        },
    )

    process_workbooks = PythonOperator(
        task_id="process_workbooks",
        python_callable=import_workbooks,
        op_kwargs={
            "bucket_name": "tc-workbooks",
            "snowflake_conn": "snowflake_tclegacy",
            "destination_schema": "PUBLIC",
            "destination_table": "WORKBOOKS",
            "num_threads": 10,
        },
        executor_config={
            "KubernetesExecutor": {
                "annotations": {
                    "iam.amazonaws.com/role": "arn:aws:iam::810110616880:role/"
                    "KubernetesAirflowProductionWorkbooksRole"
                }
            }
        },
    )

    dag << grab_workbooks >> process_workbooks
